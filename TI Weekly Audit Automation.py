import sys
sys.stdout.reconfigure(encoding='utf-8')
import os
import datetime
import openpyxl
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
import openpyxl.utils
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
import time
import shutil
from copy import copy
from openpyxl.utils import get_column_letter
import random
from collections import Counter
from selenium.webdriver.chrome.options import Options

url = "http://tstpas/TPAS/index.jsp"
user_id = "A005772"
user_pw = "A44444444444"

path_to_new_chromedriver = r'D:\本地應用程式\chromedriver-win64\chromedriver.exe'

chrome_options = Options()
chrome_options.add_experimental_option("prefs", {
    "download.default_directory": r"C:\Users\A005772\Downloads",  # 指定下載路徑
    "download.prompt_for_download": False,  # 關閉下載提示
    "download.directory_upgrade": True,
    "safebrowsing.enabled": False,  # 關閉安全瀏覽
    "safebrowsing.disable_download_protection": True,  # 嘗試關閉下載保護
    "profile.default_content_setting_values.automatic_downloads": 1  # 允許自動下載
})

s = Service(path_to_new_chromedriver)
driver = webdriver.Chrome(service=s, options=chrome_options)

# 打開網站
driver.get(url)

# 等待 mainFrame 加載完成
wait = WebDriverWait(driver, 10)
wait.until(EC.frame_to_be_available_and_switch_to_it((By.NAME, "mainFrame")))

# 定位 User ID 和 User PW 輸入框 (根據網頁元素的name等定位)
user_id_input = driver.find_element(By.NAME, "userid")
user_pw_input = driver.find_element(By.NAME, "password")

# 輸入 User ID 和 User PW
user_id_input.send_keys(user_id)
user_pw_input.send_keys(user_pw)

# 提交登錄表單
login_form = driver.find_element(By.NAME, "form1")
login_form.submit()

# 如果需要, 切換回主視窗
driver.switch_to.default_content()

# 切換到包含 'TPAS MFG Utility' 的 frame
frame = WebDriverWait(driver, 1000).until(EC.presence_of_element_located((By.XPATH, '//frame[@src="/TPAS/edu/edu_menu.jsp"]')))
driver.switch_to.frame(frame)

# 暫停 0 秒以觀察網頁操作
time.sleep(0)

# 點擊 'MFG Utility' 連結
xpath = '//a[@id="ygtvlabelel119"]'
element_to_click = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, xpath)))
element_to_click.click()

# 點擊 'MFG Use' 連結
xpath = '//a[@id="ygtvlabelel201"]'
element_to_click = WebDriverWait(driver, 0.001).until(EC.visibility_of_element_located((By.XPATH, xpath)))
element_to_click.click()

# 點擊 'Process Output' 連結
xpath = '//a[@id="ygtvlabelel216"]'
element_to_click = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, xpath)))
element_to_click.click()

# 將焦點切換回父框架
driver.switch_to.parent_frame()

# 切換到包含 'AVI Inspection Report' 的 frame
frame = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//frame[@src="/TPAS/edu/edu_main.jsp"]')))
driver.switch_to.frame(frame)

# 定位並選擇"IQC" 選項
select_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "pd_name")))
select = Select(select_element)
select.select_by_visible_text("IQC")

# 計算今天的日期
today = datetime.date.today()
today_str = today.strftime("%Y-%m-%d")

# 找到名為 "begin_date" 的輸入框並設置新的日期值（今天的日期）
begin_date_input = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.NAME, "begin_date")))
begin_date_input.clear()
begin_date_input.send_keys(today_str)

# 找到名為 "end_date" 的輸入框並設置新的日期值（今天的日期）
end_date_input = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.NAME, "end_date")))
end_date_input.clear()
end_date_input.send_keys(today_str)

# 使用更精確的 XPath 定位並選擇 "19:20" 選項
select_element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//select[@name="end_time" and option[@value="19:20"]]')))
select = Select(select_element)
select.select_by_value("19:20")

# 找到名為 "Image29" 的按鈕並點擊
submit_button = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.NAME, "Image29")))
submit_button.click()

# 暫停 10 秒以觀察網頁操作
time.sleep(0)

# 找到excel下載視窗並點擊
submit_button = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.XPATH, "//img[@name='Image13'][@src='/ArdentecW/image/icon_excel.gif']")))
submit_button.click()

# 暫停 10 秒以觀察網頁操作
time.sleep(10)

# 最後，關閉瀏覽器
driver.quit()

# 根據開始和結束時間過濾文件中的行
def filter_rows_by_time(file_path, start_time, end_time):
    # 載入文件
    wb = load_workbook(file_path)
    ws = wb.active
    time_col = 'U'
    time_col_idx = openpyxl.utils.column_index_from_string(time_col)

    rows_to_remove = []

    # 遍歷所有行
    for row_idx, row in enumerate(ws.iter_rows(min_col=time_col_idx, max_col=time_col_idx), start=1):
        if row_idx == 1:
            continue

        cell = row[0]
        cell_value = cell.value

        # 嘗試從單元格值中提取時間
        if isinstance(cell_value, datetime.datetime):
            dt = cell_value
        elif isinstance(cell_value, str):
            try:
                dt = datetime.datetime.strptime(cell_value, '%Y/%m/%d %I:%M:%S %p')
            except ValueError:
                try:
                    dt = datetime.datetime.strptime(cell_value, '%Y/%m/%d %H:%M')
                except ValueError:
                    continue
        else:
            continue

        # 判斷是否需要刪除該行
        if dt.time() < start_time or dt.time() > end_time:
            rows_to_remove.append(row_idx)

    # 刪除過濾後的行
    for row_idx in sorted(rows_to_remove, reverse=True):
        ws.delete_rows(row_idx)

    # 保存修改後的文件
    wb.save(file_path)

# 如果不存在，則在文件中創建工作表
def create_sheet_if_not_exist(file_path, sheet_name):
    wb = openpyxl.load_workbook(file_path)
    if sheet_name not in wb.sheetnames:
        control_item_index = wb.sheetnames.index("control item")
        wb.create_sheet(sheet_name, index=control_item_index + 1)
    wb.save(file_path)

def main():
    # 指定 Process_Output 和 Wafer Safety checking list 的文件路徑
    process_output_path = r"C:\Users\A005772\Downloads\Process_Output.xlsx"
    wafer_safety_checking_list_path = r"D:\Racky\工作相關\TI稽核相關 填寫路徑\QE audit result.xlsx"

    # 獲取當前日期、年份和周數
    today = datetime.date.today()
    current_year = int(str(today.year)[2:])
    current_week = today.isocalendar()[1]

    # 根據周數創建工作表
    for week in range(1, current_week + 1):
        sheet_name = f"{current_year}{week:02}"
        create_sheet_if_not_exist(wafer_safety_checking_list_path, sheet_name)

    # 設定過濾時間範圍
    start_time = datetime.time(hour=1, minute=0, second=0)
    end_time = datetime.time(hour=23, minute=0, second=0)
    # 執行過濾
    filter_rows_by_time(process_output_path, start_time, end_time)

# 程式入口點
if __name__ == "__main__":
    # 執行 main 函數
    main()

# 檔案路徑和檔名
file_path = "D:\\Racky\\工作相關\\TI稽核相關 填寫路徑\\QE audit result.xlsx"

# 開啟檔案
workbook = openpyxl.load_workbook(file_path)

# 找到control item工作表的位置
control_item_index = None
for index, sheet in enumerate(workbook.worksheets):
    if sheet.title == "control item":
        control_item_index = index
        break

if control_item_index is not None:
    # 選擇僅次於control item的工作表和往後第10個工作表
    next_sheet = workbook.worksheets[control_item_index + 1]
    tenth_sheet = workbook.worksheets[control_item_index + 10]

    # 檢查A1儲存格是否為空
    if not next_sheet['A1'].value:
        # 複製範圍A1:K33的值和格式
        for row in range(1, 34):
            for col in range(1, 12):  # 改為12，因為範圍A1:K33有11個欄位
                source_cell = tenth_sheet.cell(row=row, column=col)
                target_cell = next_sheet.cell(row=row, column=col)
                if source_cell.has_style:
                    target_cell._style = copy(source_cell._style)
                target_cell.value = source_cell.value

        # 複製欄寬和列寬
        for idx, column in enumerate(tenth_sheet.column_dimensions):
            next_sheet.column_dimensions[get_column_letter(idx + 1)].width = tenth_sheet.column_dimensions[column].width
        for idx, row in enumerate(tenth_sheet.row_dimensions):
            next_sheet.row_dimensions[idx + 1].height = tenth_sheet.row_dimensions[row].height

        # 更改索引標籤色彩
        gold_color = "DAA520"  # 接近金色,輔色4,較淺60%
        tenth_sheet.sheet_properties.tabColor = gold_color

        # 儲存檔案
        workbook.save(file_path)

        print("完成")
    else:
        print("A1錯誤")
else:
    print("未找到control item")

# 1. 打開QE audit result.xlsx
file_path = 'D:\\Racky\\工作相關\\TI稽核相關 填寫路徑\\QE audit result.xlsx'

# 2. 讀取Excel文件
workbook = openpyxl.load_workbook(file_path)
sheet = workbook.active

# 3. 讀取H列，找出儲存格內最多三種相同的值
h_col = 8  # 假設H列在第8行，如有需要請更改
h_values_list = []

for row in range(2, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=h_col).value
    if cell_value and cell_value != "Emp. No":
        h_values_list.append(cell_value)

counter_h = Counter(h_values_list)
top_three_h_values = counter_h.most_common(3)

# 4. 讀取I列，找出儲存格內最多三種相同的值
i_col = 9  # 假設I列在第9行，如有需要請更改
i_values_list = []

for row in range(2, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=i_col).value
    if cell_value and cell_value != "Lot ID":
        i_values_list.append(cell_value)

counter_i = Counter(i_values_list)
top_three_i_values = counter_i.most_common(3)

# 輸出結果
if len(top_three_h_values) <= 3 and len(top_three_i_values) <= 3:
    for i, value in enumerate(top_three_h_values):
        print(f"H_member_{i+1}: {value[0]}")
    for i, value in enumerate(top_three_i_values):
        print(f"I_member_{i+1}: {value[0]}")
else:
    print("數量錯誤")

# 新需求
# 2. 打開Process_Output.xlsx
output_file_path = 'C:\\Users\\A005772\\Downloads\\Process_Output.xlsx'

# 3. 讀取Excel文件
output_workbook = openpyxl.load_workbook(output_file_path)
output_sheet = output_workbook.active

# 4. 讀取F列的儲存格內容
f_col = 6  # 假設F列在第6行，如有需要請更改
f_values_list = []

for row in range(2, output_sheet.max_row + 1):
    cell_value = output_sheet.cell(row=row, column=f_col).value
    if cell_value and "LOT ID" not in cell_value and "TOTAL PCS" not in cell_value:
        f_values_list.append(cell_value)

# 5. 讀取W列的儲存格內容
w_col = 23  # 假設W列在第23行，如有需要請更改
w_values_list = []

for row in range(2, output_sheet.max_row + 1):
    cell_value = output_sheet.cell(row=row, column=w_col).value
    #修改前if cell_value and "Emp. No" not in cell_value:
    if str(cell_value) and "Emp. No" not in str(cell_value):
        w_values_list.append(cell_value)

# 隨機抽取最多3個替換I_member並且不重複，並且抽到的值不能包含"LOT ID"或"TOTAL PCS"
replace_i_members = []
while len(replace_i_members) < min(3, len(top_three_i_values)):
    candidate = random.choice(f_values_list)
    if candidate not in replace_i_members and "LOT ID" not in candidate and "TOTAL PCS" not in candidate:
        replace_i_members.append(candidate)

for i, value in enumerate(replace_i_members):
    print(f"新I_member_{i+1}: {value}")

# 更新QE audit result.xlsx中的I列變數
for row in range(2, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=i_col).value
    if cell_value and cell_value != "Lot ID" and cell_value in [x[0] for x in top_three_i_values]:
        sheet.cell(row=row, column=i_col).value = replace_i_members[top_three_i_values.index((cell_value, counter_i[cell_value]))]

# 從未被選為I_member的值中隨機抽取最多3個替換H_member並且不重複，並且抽到的值不能包含"TA"
remaining_values = list(set(w_values_list) - set(replace_i_members))
replace_h_members = []

while len(replace_h_members) < min(3, len(remaining_values)):
    candidate = random.choice(remaining_values)
    if candidate is not None and "TA" not in candidate and candidate not in replace_h_members:
    #if candidate is not None and "TA" not in str(candidate) and str(candidate) not in [str(m) for m in replace_h_members]:
        replace_h_members.append(candidate)

for i, value in enumerate(replace_h_members):
    print(f"新H_member_{i+1}: {value}")

# 更新QE audit result.xlsx中的H列變數
for row in range(2, sheet.max_row + 1):
    cell_value = sheet.cell(row=row, column=h_col).value
    if cell_value and cell_value != "Emp. No" and cell_value in [x[0] for x in top_three_h_values]:
        sheet.cell(row=row, column=h_col).value = replace_h_members[top_three_h_values.index((cell_value, counter_h[cell_value]))]

# 儲存修改後的檔案
workbook.save(file_path)

# 檔案路徑
file_path = "D:\\Racky\\工作相關\\TI稽核相關 填寫路徑\\QE audit result.xlsx"

# 獲取當前日期、年份和周數
today = datetime.date.today()
current_year = int(str(today.year)[2:])
current_week = today.isocalendar()[1]

# 複製檔案並重新命名
new_file_name = f"QE audit result 2024 - W{current_year}{current_week}.xlsx"
new_file_path = os.path.join(os.path.dirname(file_path), new_file_name)

# 如果檔名已存在，刪除原有檔案
if os.path.exists(new_file_path):
    os.remove(new_file_path)

# 複製檔案
shutil.copy(file_path, new_file_path)

print(f"程式已執行完畢：{new_file_name}")
